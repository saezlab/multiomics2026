#!/usr/bin/env python3
"""Extract data sheets from supplementary xlsx files to TSV.

Reads the second sheet (data) from each relevant EV table and writes
it as a tab-separated file into the project data/ directory.
"""

from pathlib import Path
import gzip

import openpyxl
import csv

EV_DIR = Path(__file__).parent.parent / "project-prep" / "ev-tables"
DATA_DIR = Path(__file__).parent / "data"

# Mapping: (xlsx filename suffix, output path, description)
TABLES = [
    ("MOESM4_ESM.xlsx", "differential/diff_expr_all.tsv.gz",
     "Differential expression results, all omics"),
    ("MOESM6_ESM.xlsx", "differential/tf_activities_ev3.tsv",
     "TF activity enrichment results (from EV3 only, no kinases)"),
    ("MOESM7_ESM.xlsx", "network/paper_edges.tsv",
     "Network edge table from paper"),
    ("MOESM8_ESM.xlsx", "network/paper_nodes.tsv",
     "Network node table from paper"),
    ("MOESM9_ESM.xlsx", "imaging/col1_timecourse.tsv",
     "COL1 imaging time course data"),
]


def find_xlsx(suffix: str) -> Path:
    matches = list(EV_DIR.glob(f"*{suffix}"))
    if not matches:
        raise FileNotFoundError(f"No file matching *{suffix} in {EV_DIR}")
    return matches[0]


def extract_sheet(xlsx_path: Path, output_path: Path):
    """Read second sheet from xlsx and write as TSV."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[1]  # second sheet is data
    ws = wb[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)

    opener = gzip.open if output_path.suffix == ".gz" else open
    with opener(output_path, "wt", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            # Clean newlines within cells (e.g. "Kinase/\nphosphatase")
            cleaned = [
                str(v).replace("\n", " ").replace("\r", "") if v is not None else ""
                for v in row
            ]
            writer.writerow(cleaned)

    wb.close()


def main():
    for suffix, rel_path, desc in TABLES:
        xlsx_path = find_xlsx(suffix)
        output_path = DATA_DIR / rel_path
        print(f"  {xlsx_path.name} -> {rel_path} ... ", end="", flush=True)
        extract_sheet(xlsx_path, output_path)
        # Count lines
        n_lines = sum(1 for _ in opener(output_path, "rt")) - 1  # minus header
        print(f"{n_lines} rows")

    print("\nDone. All data extracted to", DATA_DIR)


if __name__ == "__main__":
    main()
